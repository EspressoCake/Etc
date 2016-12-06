__author__ = 'EspressoCake'

#These import statements are excessive, and I am aware of it.
import warnings
warnings.simplefilter("ignore", UserWarning)
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import csv
import xlsxwriter
import string
import pyperclip
import webbrowser

#Converts character input into a numerical value for manipulation.
def convertValues(letter):
    num = 0
    for c in letter:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return(num)

def workbookParser():
    workbook = input("\nWhat/where is the file you need to access? \n").replace('\\', '/')
    outputWorkbook = input("\nWhat folder would you like the modified file to save to? \n").replace('\\', '/')
    columnNumber = convertValues(input("\nWhat column letter are the IP hosts listed under?: \n"))
    wb = load_workbook(workbook)
    ws1 = wb.get_sheet_by_name('Host Assessment')

    testvariable = chr((columnNumber - 1) + ord('A'))

    print("\n\nPROCESSING...\n")
    for i in range(3, 500):
        if (ws1.cell(row=i, column=1)).value != None:
            if (ws1.cell(row=i, column=7)).value == None or (ws1.cell(row=i, column=7)).value == '':
                (ws1.cell(row=i, column=7)).value = 'N/A'
            else:
                continue
        else:
            break

    for i in range(3, 500):
        if (ws1.cell(row=i, column=1)).value != None:
            if (ws1.cell(row=i, column=8)).value == None or (ws1.cell(row=i, column=8)).value == '':
                (ws1.cell(row=i, column=8)).value = 'N/A'
            else:
                continue
        else:
            break

    for i in range(3, 500):
        if (ws1.cell(row=i, column=1)).value != None:
            if (ws1.cell(row=i, column=12)).value == None or (ws1.cell(row=i, column=12)).value == '':
                (ws1.cell(row=i, column=12)).value = 'N/A'
            else:
                continue
        else:
            break

    for i in range(3, 500):
        if (ws1.cell(row=i, column=1)).value != None:
            if (ws1.cell(row=i, column=13)).value == None or (ws1.cell(row=i, column=13)).value == '':
                (ws1.cell(row=i, column=13)).value = 'No'
            else:
                continue
        else:
            break

    for i in range(1, 500):
        if (ws1.cell(row=i, column=columnNumber)).value == 'Afftected Instances':
            (ws1.cell(row=i, column=columnNumber)).value = 'Affected Instances'
        elif (ws1.cell(row=i, column=columnNumber)).value == None or (ws1.cell(row=i, column=columnNumber)).value == 'Affected Instances':
            continue
        else:
            string = ''
            splitText = ws1.cell(row=i, column=columnNumber).value.split(';')
            List = sorted(splitText, key=lambda x:tuple(map(int, x.split('.'))))
            for item in List:
                string += str(item) + '\n'
                ws1.cell(row=i, column=columnNumber).value = string
            ws1.column_dimensions[testvariable].width = 30.0
    wb.save(outputWorkbook + "/modifiedoutput.xlsx")

    print(str("\nSUCCESS!\nNew file has been saved in " + outputWorkbook + "/modifiedoutput.xlsx.").replace('//', '/'))
    pyperclip.copy(outputWorkbook + "/modifiedoutput.xlsx")
    webbrowser.open(outputWorkbook)

if __name__ == "__main__":
    workbookParser()
